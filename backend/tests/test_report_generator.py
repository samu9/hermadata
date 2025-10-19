from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook

from hermadata.constants import EntryType, ExitType
from hermadata.reports.report_generator import (
    AdopterVariables,
    AnimalVariables,
    ReportAdoptionVariables,
    ReportAnimalEntryVariables,
    ReportChipAssignmentVariables,
    ReportFormat,
    ReportGenerator,
    ReportVariationVariables,
)
from hermadata.repositories.animal.models import (
    AnimalDaysItem,
    AnimalDaysQuery,
    AnimalDaysResult,
    AnimalEntriesItem,
    AnimalEntriesQuery,
    AnimalReportResult,
)
from hermadata.settings import Settings


def test_animal_entry_report(
    report_generator: ReportGenerator, test_settings: Settings
):
    variables = ReportAnimalEntryVariables(
        city="Test",
        animal_name="Gino",
        animal_type="Gatto",
        entry_date=date(2020, 2, 1),
    )
    pdf = report_generator.build_animal_entry_report(variables)

    with open(
        test_settings.storage.disk.base_path + "/generated.pdf", "wb"
    ) as fp:
        fp.write(pdf)

    assert pdf


def test_chip_assignment_report(
    report_generator: ReportGenerator, test_settings
):
    variables = ReportChipAssignmentVariables(
        animal_name="Gino",
        chip_code="111.111.111.111.111",
        assignment_date=datetime.now().date(),
    )
    pdf = report_generator.build_chip_assignment_report(variables)

    with open(
        test_settings.storage.disk.base_path + "/generated.pdf", "wb"
    ) as fp:
        fp.write(pdf)

    assert pdf


def test_animal_days_report(report_generator: ReportGenerator, test_settings):
    data = AnimalDaysResult(
        items=[
            AnimalDaysItem(
                animal_name="Test",
                animal_chip_code="123.123.123.123.123",
                animal_days=5,
            )
        ],
        total_days=5,
    )
    query = AnimalDaysQuery(
        from_date=date(2023, 1, 1),
        to_date=date(2023, 2, 1),
        city_code="H501",
    )
    filename, report = report_generator.generate_animal_days_count_report(
        query, data
    )

    assert isinstance(filename, str)
    assert (
        filename == f"giorni_cane{query.from_date.strftime('%Y-%m-%d')}"
        f"_{query.to_date.strftime('%Y-%m-%d')}.xlsx"
    )
    with open(
        test_settings.storage.disk.base_path + "/generated.xls", "wb"
    ) as fp:
        fp.write(report)

    assert isinstance(report, bytes)


def test_animal_entries_report(empty_db, report_generator: ReportGenerator):
    query = AnimalEntriesQuery(
        from_date=date(2024, 1, 1),
        to_date=date(2024, 12, 31),
        entry_type=EntryType.rescue,
        city_code="H501",
    )
    data = AnimalReportResult[AnimalEntriesItem](
        items=[
            AnimalEntriesItem(
                animal_race="Cane",
                animal_name="Dino",
                animal_chip_code="123.123.123.123.123",
                animal_birth_date=date(2020, 1, 1),
                animal_sex="M",
                entry_city="Montecatini Terme",
                entry_date=date(2024, 1, 2),
                entry_type=EntryType.rescue,
            )
        ],
        total=1,
    )
    filename, report = report_generator.generate_animal_entries_report(
        query, data, ReportFormat.excel
    )

    assert (
        filename == f"ingressi_{query.from_date.strftime('%Y-%m-%d')}"
        f"_{query.to_date.strftime('%Y-%m-%d')}.xlsx"
    )
    file_ = BytesIO(report)
    wb = load_workbook(file_)

    sheet = wb[wb.sheetnames[0]]

    rows = list(sheet.rows)
    assert len(rows) == 2

    assert rows[1][0].value == "Cane"
    assert rows[1][1].value == "Dino"
    assert rows[1][4].value == "M"


def test_variation_report(report_generator: ReportGenerator):
    variables = ReportVariationVariables(
        animal=AnimalVariables(
            name="Gino",
            chip_code="111.111.111.111.111",
            fur_type="lungo",
            age=12,
            breed="Chihuahua",
            fur_color="binaco",
            origin_city="Montecatini terme",
            sex="maschio",
            entry_date=datetime.now().date(),
        ),
        variation_date=datetime.now().date(),
        adopter=AdopterVariables(
            name="Mario",
            surname="Rossi",
            residence_city="Montecatini Terme",
            birth_city="Roma",
            birth_date=date(1979, 2, 3),
        ),
        variation_type=ExitType.custody,
    )
    pdf = report_generator.build_variation_report(variables)

    assert pdf


def test_adoption_report(report_generator: ReportGenerator):
    variables = ReportAdoptionVariables(
        animal=AnimalVariables(
            name="Gino",
            chip_code="111.111.111.111.111",
            fur_type="lungo",
            age=12,
            breed="Chihuahua",
            fur_color="binaco",
            origin_city="Montecatini terme",
            sex="maschio",
            entry_date=datetime.now().date() - timedelta(days=30),
        ),
        adopter=AdopterVariables(
            name="Mario",
            surname="Rossi",
            residence_city="Montecatini Terme",
            birth_city="Roma",
            birth_date=date(1979, 2, 3),
            residence_address="Via test",
            phone="123456789",
        ),
        exit_date=datetime.now().date(),
        notes="Test",
        location_address="Via Prova",
        location_city="LUCCA",
        location_province="LU",
    )

    pdf = report_generator.build_adoption_report(variables)

    assert pdf


def test_models():
    v = AnimalVariables(
        name="test",
        age=4,
        breed="test",
        chip_code="test",
        entry_date=date(2022, 4, 5),
        fur_color="test",
        fur_type="test",
        origin_city="test",
        sex="M",
    )

    assert v.age == 4

    dump = v.model_dump()

    assert dump["entry_date"] == "05/04/2022"
    assert dump["age"] == 4

    r = ReportVariationVariables(
        animal=v,
        variation_type=ExitType.adoption,
        variation_date=datetime.now().date(),
    )

    dump = r.model_dump()

    assert dump["animal"]["age"] == 4
