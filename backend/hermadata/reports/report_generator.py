from datetime import date, datetime
from enum import Enum
from io import BytesIO
from typing import Annotated

import openpyxl
from jinja2 import Environment
from pydantic import (
    BaseModel,
    Field,
    PlainSerializer,
    field_serializer,
    field_validator,
)
from weasyprint import CSS, HTML

from hermadata.constants import (
    ENTRY_TYPE_LABELS,
    EXIT_TYPE_LABELS,
    FUR_LABELS,
    AnimalFur,
    ExitType,
)
from hermadata.repositories.animal.models import (
    AnimalDaysQuery,
    AnimalDaysResult,
    AnimalEntriesItem,
    AnimalEntriesQuery,
    AnimalExitsItem,
    AnimalExitsQuery,
    AnimalReportResult,
)

ReportDate = Annotated[
    date, PlainSerializer(lambda x: x.strftime("%d/%m/%Y"), return_type=str)
]


def transform_date_to_string(raw: date) -> str:
    return raw.strftime("%d/%m/%Y")


class ReportFormat(Enum):
    pdf = "application/pdf"
    excel = "xls"


DEFAULT_EXTENSIONS: dict[ReportFormat, str] = {ReportFormat.excel: "xlsx"}


class BaseVariables(BaseModel):
    pass


class ReportDefaultVariables(BaseVariables):
    day: ReportDate = Field(default_factory=lambda: datetime.now().date())
    title: str

    @field_serializer("day")
    def serialize_day(self, day: date):
        return day.strftime("%Y-%m-%d")


class ReportAnimalEntryVariables(ReportDefaultVariables):
    title: str = "INGRESSO ANIMALE"
    city: str
    animal_type: str
    animal_name: str | None = None
    entry_date: ReportDate


class ReportChipAssignmentVariables(ReportDefaultVariables):
    title: str = "ASSEGNAMENTO CHIP"
    animal_name: str | None = None
    chip_code: str
    assignment_date: ReportDate


class AdopterVariables(BaseModel):
    name: str
    surname: str
    residence_city: str
    residence_address: str | None = ""
    birth_city: str
    birth_date: ReportDate


class AnimalVariables(BaseVariables):
    name: str | None = ""
    chip_code: str
    breed: str | None = ""
    sex: str | None = ""
    age: int | None = ""
    fur_type: str | None = ""
    fur_color: str | int | None = ""
    origin_city: str
    entry_date: ReportDate

    @field_validator("sex", mode="before")
    def validate_sex(value: int | str):
        if isinstance(value, int):
            value = "M" if value == 0 else "F"
        return value

    @field_validator("fur_type", mode="before")
    def validate_fur(value: int | str):
        if isinstance(value, int):
            value = FUR_LABELS[AnimalFur(value)]
        return value


class ReportVariationVariables(ReportDefaultVariables):
    title: str = "VARIAZIONE"
    animal: AnimalVariables
    adopter: AdopterVariables | None = None

    variation_type: ExitType  # scomparso, deceduto, stato ceduto
    variation_date: ReportDate


class ReportAdoptionVariables(ReportDefaultVariables):
    title: str = "DOCUMENTO DI ADOZIONE"
    animal_name: str
    chip_code: str
    exit_date: ReportDate


class ReportCustodyVariables(ReportDefaultVariables):
    title: str = "DOCUMENTO DI AFFIDO"
    animal_name: str
    chip_code: str
    exit_date: ReportDate


class ReportGenerator:
    def __init__(self, jinja_env: Environment) -> None:
        self.jinja_env = jinja_env

    def _build_template(
        self, filename: str, variables: ReportDefaultVariables
    ) -> bytes:
        template = self.jinja_env.get_template(filename)

        rendered_html = template.render(**variables.model_dump())

        with open("output.html", "w") as fp:
            fp.write(rendered_html)

        target = BytesIO()

        HTML(string=rendered_html).write_pdf(
            target=target,
            stylesheets=[
                CSS(filename="hermadata/reports/static/tailwind.css")
            ],
        )

        return target.getvalue()

    def build_animal_entry_report(
        self, variables: ReportAnimalEntryVariables
    ) -> bytes:
        return self._build_template("animal_entry.jinja", variables)

    def build_chip_assignment_report(
        self, variables: ReportChipAssignmentVariables
    ) -> bytes:
        return self._build_template("chip_assignment.jinja", variables)

    def build_adoption_report(
        self, variables: ReportAdoptionVariables
    ) -> bytes:
        return self._build_template("adoption.jinja", variables)

    def build_custody_report(self, variables: ReportCustodyVariables) -> bytes:
        return self._build_template("custody.jinja", variables)

    def build_variation_report(
        self, variables: ReportVariationVariables
    ) -> bytes:
        return self._build_template("variation.jinja", variables)

    def generate_animal_days_count_report(
        self,
        query: AnimalDaysQuery,
        data: AnimalDaysResult,
        format: ReportFormat = ReportFormat.excel,
    ) -> tuple[str, bytes]:
        if format != ReportFormat.excel:
            raise Exception("format not supported")

        wb = openpyxl.Workbook()

        ws = wb.active
        ws.append(["Nome", "Chip", "Giorni"])
        for d in data.items:
            ws.append([d.animal_name, d.animal_chip_code, d.animal_days])

        ws.append([])
        ws.append(["Totale", "", data.total_days])

        filename = (
            f"giorni_cane{query.from_date.strftime('%Y-%m-%d')}"
            f"_{query.to_date.strftime('%Y-%m-%d')}"
            f".{DEFAULT_EXTENSIONS[format]}"
        )

        fp = BytesIO()
        wb.save(fp)

        bytes_data = fp.getvalue()

        return filename, bytes_data

    def generate_animal_entries_report(
        self,
        query: AnimalEntriesQuery,
        data: AnimalReportResult[AnimalEntriesItem],
        format: ReportFormat = ReportFormat.excel,
    ) -> tuple[str, bytes]:
        if format != ReportFormat.excel:
            raise Exception("format not supported")

        wb = openpyxl.Workbook()

        ws = wb.active
        ws.append(
            [
                "Tipo",
                "Nome",
                "Chip",
                "Data nascita",
                "Sesso",
                "Data ingresso",
                "Tipo Ingresso",
                "Comune",
            ]
        )
        for d in data.items:
            ws.append(
                [
                    d.animal_race,
                    d.animal_name,
                    d.animal_chip_code,
                    d.animal_birth_date,
                    d.animal_sex,
                    d.entry_date,
                    ENTRY_TYPE_LABELS[d.entry_type],
                    d.entry_city,
                ]
            )

        filename = (
            f"ingressi_{query.from_date.strftime('%Y-%m-%d')}"
            f"_{query.to_date.strftime('%Y-%m-%d')}"
            f".{DEFAULT_EXTENSIONS[format]}"
        )

        fp = BytesIO()
        wb.save(fp)

        bytes_data = fp.getvalue()

        return filename, bytes_data

    def generate_animal_exits_report(
        self,
        query: AnimalExitsQuery,
        data: AnimalReportResult[AnimalExitsItem],
        format: ReportFormat = ReportFormat.excel,
    ) -> tuple[str, bytes]:
        if format != ReportFormat.excel:
            raise Exception("format not supported")

        wb = openpyxl.Workbook()

        ws = wb.active
        ws.append(
            [
                "Tipo",
                "Nome",
                "Chip",
                "Data nascita",
                "Sesso",
                "Data uscita",
                "Tipo uscita",
            ]
        )
        for d in data.items:
            ws.append(
                [
                    d.animal_race,
                    d.animal_name,
                    d.animal_chip_code,
                    d.animal_birth_date,
                    d.animal_sex,
                    d.exit_date,
                    EXIT_TYPE_LABELS[d.exit_type],
                ]
            )

        filename = (
            f"uscite_{query.from_date.strftime('%Y-%m-%d')}"
            f"_{query.to_date.strftime('%Y-%m-%d')}"
            f".{DEFAULT_EXTENSIONS[format]}"
        )

        fp = BytesIO()
        wb.save(fp)

        bytes_data = fp.getvalue()

        return filename, bytes_data
